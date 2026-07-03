import openpyxl
import os

excel_path = '/Users/harald2018/Library/CloudStorage/GoogleDrive-harald.pilch@e-mobility-brands.com/Meine Ablage/Blaupunkt EV/Preislisten/2024_reseller_OTC_Range_RRP.xlsx'

print("Checking file existence...")
if not os.path.exists(excel_path):
    print(f"Error: File does not exist at {excel_path}")
    exit(1)

print("Opening workbook...")
wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
print("Sheets in workbook:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    print(f"\n--- Sheet: {sheet_name} ---")
    sheet = wb[sheet_name]
    # Read the first 10 rows
    rows = list(sheet.iter_rows(max_row=10, values_only=True))
    if not rows:
        print("Empty sheet.")
        continue
    for i, row in enumerate(rows):
        # Print non-empty rows or just print first few
        print(f"Row {i+1}: {row[:15]}") # Print first 15 columns of each row
